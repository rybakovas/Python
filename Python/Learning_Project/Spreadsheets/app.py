import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('items.xlsx')
sheet = wb['Plan1']
cell = sheet['a1']
cell = sheet.cell(1, 1)
print(cell.value)
print(sheet.max_row)

for row in range(1, sheet.max_row + 1):
    print(row)
# ================================================
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    print(cell.value)
# ================================================

wb = xl.load_workbook('items.xlsx')
sheet = wb['Plan1']
cell = sheet['a1']

for row in range(2, sheet.max_row):
    cell = sheet.cell(row, 3)
    correct_price = float(cell.value) * 0.9
    correct_price_cell = sheet.cell(row, 5)
    correct_price_cell.value = correct_price

wb.save('items2.xlsx')

# ================================================


def process_workbook(filename):

    wb = xl.load_workbook(filename)
    sheet = wb['Plan1']

    for row in range(2, sheet.max_row):
        cell = sheet.cell(row, 3)
        correct_price = float(cell.value) * 0.9
        correct_price_cell = sheet.cell(row, 5)
        correct_price_cell.value = correct_price

    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=5, max_col=5)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'f2')

    wb.save('items2.xlsx')



# By Victor Rybakovas Sep 2019 - http://bit.ly/linkedin-victor

